#!/usr/bin/python
#_*_ coding:utf-8 _*_

import openpyxl
from openpyxl import Workbook
#读取Excel需要load_workbook
from openpyxl import load_workbook

wb = Workbook()
ws=wb.active
#更改默认名称sheet
ws.title="s1"

#第2行第2列赋值
#ws.cell(row=2,column=2,value=20)

#更改Tab颜色
#ws.sheet_properties.tabColor = "1072BA"

#给单元格赋值
def fun1():
    ws["A1"]="姓名"
    ws["B1"]="年龄"
    ws["C1"]="工作"

#指定行列给单元格赋值
def fun2():
    v=0
    for i in range(1,10):
        for j in range(1,10):
            ws.cell(row=i,column=j,value=v)
            v+=1


def fun3():
    column_title = ["FirstName", "LastName"]
    rows=[column_title,["Tarou", "Tanaka"],["Tarou", "Suzuki"],["Tarou", "Uchiayama"],]
    for row in rows:
        ws.append(row)

#换行
def fun4():
    ws["A1"]="A\nB\nC"
    ws["A1"].alignment=openpyxl.styles.Alignment(wrapText=True)

#合并单元格
def fun5():
    ws.merge_cells("A1:D2")
    ws["A1"]="合并单元格"


if __name__=="__main__":
    #fun1()
    #fun2()
    #fun3()
    #fun4()
    fun5()
    wb.save("user.xlsx")
